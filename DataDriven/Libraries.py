import requests
import json
import jsonpath
import openpyxl

class Common:
    def __init__(self, FileNamePath, SheetName):
        global workbook
        global sheet
        workbook = openpyxl.load_workbook(FileNamePath)
        sheet = workbook[SheetName]

    def fetch_row_count(self):
        rows = sheet.max_row
        return rows
    
    def fetch_col_count(self):
        col = sheet.max_column
        return col
    
    def fetch_key_names(self):
        c = sheet.max_column
        li = []
        for i in range(1, c+1):
            cell = sheet.cell(row=1, column=i)
            li.insert(i, cell.value)

        return li
    
    def update_request_with_data(self, rowNumber, jsonRequest, keyList):
        c = sheet.max_column
        for i in range(1, c+1):
            cell = sheet.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]] = cell.value

        return jsonRequest