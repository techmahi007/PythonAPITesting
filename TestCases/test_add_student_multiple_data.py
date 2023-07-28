import requests
import json
import jsonpath
import openpyxl

def test_add_multiple_students():
    # API
    api_url = "https://thetestingworldapi.com/api/studentsDetails"
    file = open("D:/Training_Programs/Python_Automation_July_2023/API_Testing_Python_Pytest_Allure/TestCases/AddStudent.json", "r")
    request_json = json.loads(file.read())
    
    # Excel
    workbook = openpyxl.load_workbook("C:/Users/mahimna.bhuskute/Downloads/Book.xlsx")
    sheet = workbook['Sheet1']
    rows = sheet.max_row

    for i in range(2, rows+1):
        cell_first_name = sheet.cell(row=i, column=1)
        cell_mid_name = sheet.cell(row=i, column=2)
        cell_last_name = sheet.cell(row=i, column=3)
        cell_dob = sheet.cell(row=i, column=4)

        request_json['first_name'] = cell_first_name.value
        request_json['middle_name'] = cell_mid_name.value
        request_json['last_name'] = cell_last_name.value
        request_json['date_of_birth'] = cell_dob.value

        response = requests.post(api_url, request_json)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201